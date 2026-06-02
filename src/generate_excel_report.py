import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

CPI_FILE = "data/cleaned/cleaned_cpi_ontario.csv"
LABOUR_FILE = "data/cleaned/cleaned_labour_ontario.csv"
REPORT_FILE = "reports/ontario_economic_report.xlsx"
CHARTS_FOLDER = "charts"


def auto_adjust_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 3


def style_header(ws):
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def create_excel_report():
    os.makedirs("reports", exist_ok=True)

    cpi_df = pd.read_csv(CPI_FILE)
    labour_df = pd.read_csv(LABOUR_FILE)

    # Yearly CPI average by category
    yearly_cpi = (
        cpi_df
        .groupby(["Year", "Category"])["CPI_Value"]
        .mean()
        .reset_index()
    )

    cpi_pivot = yearly_cpi.pivot(
        index="Year",
        columns="Category",
        values="CPI_Value"
    ).reset_index()

    # CPI change from 2019 to 2025
    cpi_change = []

    for category in yearly_cpi["Category"].unique():
        category_data = yearly_cpi[yearly_cpi["Category"] == category]

        value_2019 = category_data[category_data["Year"] == 2019]["CPI_Value"].mean()
        value_2025 = category_data[category_data["Year"] == 2025]["CPI_Value"].mean()

        if pd.notna(value_2019) and pd.notna(value_2025):
            percent_change = ((value_2025 - value_2019) / value_2019) * 100

            cpi_change.append({
                "Category": category,
                "Average CPI 2019": round(value_2019, 2),
                "Average CPI 2025": round(value_2025, 2),
                "Percent Change 2019-2025": round(percent_change, 2)
            })

    cpi_change_df = pd.DataFrame(cpi_change)

    # Labour yearly average
    yearly_labour = (
        labour_df
        .groupby(["Year", "Indicator"])["Value"]
        .mean()
        .reset_index()
    )

    labour_pivot = yearly_labour.pivot(
        index="Year",
        columns="Indicator",
        values="Value"
    ).reset_index()

    # Summary insights
    all_items = cpi_change_df[cpi_change_df["Category"] == "All-items"]
    food = cpi_change_df[cpi_change_df["Category"] == "Food"]
    shelter = cpi_change_df[cpi_change_df["Category"] == "Shelter"]

    summary_data = {
        "Metric": [
            "Project Period",
            "Province",
            "Main CPI Category",
            "All-items CPI Change 2019-2025 (%)",
            "Food CPI Change 2019-2025 (%)",
            "Shelter CPI Change 2019-2025 (%)",
            "Highest CPI Growth Category",
            "Lowest CPI Growth Category"
        ],
        "Value": [
            "2019 to 2025",
            "Ontario",
            "All-items, Food, Shelter, Transportation, Energy",
            all_items["Percent Change 2019-2025"].iloc[0] if not all_items.empty else "N/A",
            food["Percent Change 2019-2025"].iloc[0] if not food.empty else "N/A",
            shelter["Percent Change 2019-2025"].iloc[0] if not shelter.empty else "N/A",
            cpi_change_df.sort_values("Percent Change 2019-2025", ascending=False).iloc[0]["Category"],
            cpi_change_df.sort_values("Percent Change 2019-2025", ascending=True).iloc[0]["Category"]
        ]
    }

    summary_df = pd.DataFrame(summary_data)

    # Write Excel report
    with pd.ExcelWriter(REPORT_FILE, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        cpi_pivot.to_excel(writer, sheet_name="Yearly CPI", index=False)
        cpi_change_df.to_excel(writer, sheet_name="CPI Change", index=False)
        labour_pivot.to_excel(writer, sheet_name="Labour Market", index=False)

    # Open workbook to style and insert charts
    wb = load_workbook(REPORT_FILE)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        style_header(ws)
        auto_adjust_columns(ws)
        ws.freeze_panes = "A2"

    # Add dashboard sheet
    ws = wb.create_sheet("Dashboard")

    ws["A1"] = "Ontario Cost of Living & Job Market Analysis"
    ws["A1"].font = Font(size=18, bold=True)

    ws["A3"] = "This dashboard summarizes CPI, cost-of-living categories, employment, and unemployment trends in Ontario from 2019 to 2025."
    ws["A3"].alignment = Alignment(wrap_text=True)

    chart_files = [
        ("ontario_cpi_trend.png", "A5"),
        ("yearly_cpi_comparison.png", "J5"),
        ("unemployment_rate_trend.png", "A25"),
        ("employment_unemployment_trend.png", "J25")
    ]

    for chart_file, cell_position in chart_files:
        chart_path = os.path.join(CHARTS_FOLDER, chart_file)

        if os.path.exists(chart_path):
            img = Image(chart_path)
            img.width = 600
            img.height = 300
            ws.add_image(img, cell_position)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["J"].width = 30

    wb.save(REPORT_FILE)

    print("Excel report created successfully.")
    print(f"Saved to: {REPORT_FILE}")


if __name__ == "__main__":
    create_excel_report()